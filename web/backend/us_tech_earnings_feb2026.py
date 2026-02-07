#!/usr/bin/env python3
"""
生成美国科技公司财报Excel文件
2026年2月第一周（1月底-2月初发布）
"""

import subprocess
import sys

# 确保安装 openpyxl
subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "美国科技公司财报"

# 定义样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = [
    "公司名称", "股票代码", "财报发布日期", "财报季度",
    "营收(亿美元)", "营收同比增长", "每股收益(EPS)",
    "是否超预期", "股价反应", "主要亮点"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 财报数据
earnings_data = [
    {
        "公司名称": "苹果 (Apple)",
        "股票代码": "AAPL",
        "财报发布日期": "2026-01-30",
        "财报季度": "FY2026 Q1",
        "营收(亿美元)": 1438,
        "营收同比增长": "+16%",
        "每股收益(EPS)": "$2.84",
        "是否超预期": "✓ 超预期",
        "股价反应": "↑ 上涨约7%",
        "主要亮点": "iPhone销售创纪录；服务收入增长14%；活跃设备超25亿"
    },
    {
        "公司名称": "微软 (Microsoft)",
        "股票代码": "MSFT",
        "财报发布日期": "2026-01-29",
        "财报季度": "FY2026 Q2",
        "营收(亿美元)": 812.7,
        "营收同比增长": "+12%",
        "每股收益(EPS)": "$4.14",
        "是否超预期": "✓ 超预期",
        "股价反应": "↓ 下跌约10%",
        "主要亮点": "云收入首超500亿；Azure增长40%；净利润增60%至385亿"
    },
    {
        "公司名称": "亚马逊 (Amazon)",
        "股票代码": "AMZN",
        "财报发布日期": "2026-02-06",
        "财报季度": "2025 Q4",
        "营收(亿美元)": 2134,
        "营收同比增长": "+12%",
        "每股收益(EPS)": "$1.95",
        "是否超预期": "≈ 基本持平",
        "股价反应": "↓ 下跌约11%",
        "主要亮点": "AWS收入356亿(+24%)；AI资本支出计划引发担忧"
    },
    {
        "公司名称": "Meta",
        "股票代码": "META",
        "财报发布日期": "2026-01-29",
        "财报季度": "2025 Q4",
        "营收(亿美元)": 598.5,
        "营收同比增长": "+24%",
        "每股收益(EPS)": "$8.88",
        "是否超预期": "✓ 超预期",
        "股价反应": "↑ 上涨约10%",
        "主要亮点": "日活用户35.8亿(+7%)；计划2026年AI投资1150-1350亿"
    },
    {
        "公司名称": "谷歌 (Alphabet)",
        "股票代码": "GOOGL",
        "财报发布日期": "2026-02-04",
        "财报季度": "2025 Q4",
        "营收(亿美元)": 965,
        "营收同比增长": "+12%",
        "每股收益(EPS)": "$2.15",
        "是否超预期": "✓ 超预期",
        "股价反应": "↓ 下跌",
        "主要亮点": "利润345亿；云收入增长48%；2026年AI资本支出指引过高"
    },
    {
        "公司名称": "高通 (Qualcomm)",
        "股票代码": "QCOM",
        "财报发布日期": "2026-02-04",
        "财报季度": "FY2026 Q1",
        "营收(亿美元)": 123,
        "营收同比增长": "+5%",
        "每股收益(EPS)": "$3.50",
        "是否超预期": "✓ 超预期",
        "股价反应": "→ 持平",
        "主要亮点": "超出112亿预期；Q2指引102-110亿；手机内存供应存隐忧"
    },
    {
        "公司名称": "Snap",
        "股票代码": "SNAP",
        "财报发布日期": "2026-02-04",
        "财报季度": "2025 Q4",
        "营收(亿美元)": 15.6,
        "营收同比增长": "+14%",
        "每股收益(EPS)": "$0.16",
        "是否超预期": "✓ 超预期",
        "股价反应": "↓ 创新低$5.87",
        "主要亮点": "期权市场预期波动±12.5%；2026年前景不明朗"
    },
    {
        "公司名称": "Pinterest",
        "股票代码": "PINS",
        "财报发布日期": "2026-02-05",
        "财报季度": "2025 Q4",
        "营收(亿美元)": 13.32,
        "营收同比增长": "+18%",
        "每股收益(EPS)": "$1.15",
        "是否超预期": "✓ 超预期",
        "股价反应": "↓ 52周新低$20.12",
        "主要亮点": "全球重组计划；预计产生3500-4500万重组费用"
    },
]

# 填充数据
for row_idx, data in enumerate(earnings_data, 2):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=data[header])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 根据股价反应设置背景色
        if header == "股价反应":
            if "↑" in str(data[header]):
                cell.fill = positive_fill
            elif "↓" in str(data[header]):
                cell.fill = negative_fill

# 调整列宽
column_widths = [18, 12, 16, 14, 16, 14, 16, 14, 18, 50]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 设置行高
for row in range(1, len(earnings_data) + 2):
    ws.row_dimensions[row].height = 35

# 添加汇总信息
summary_row = len(earnings_data) + 3
ws.cell(row=summary_row, column=1, value="📊 数据汇总").font = Font(bold=True, size=12)
ws.cell(row=summary_row + 1, column=1, value="统计时间：2026年1月29日 - 2026年2月7日")
ws.cell(row=summary_row + 2, column=1, value="数据来源：各公司官方财报、华尔街分析师报告")
ws.cell(row=summary_row + 3, column=1, value=f"共收录 {len(earnings_data)} 家科技公司财报")

# 保存文件
output_path = "/Users/fangyudong/PycharmProjects/claude-agent-sdk-python/web/backend/美国科技公司财报_2026年2月.xlsx"
wb.save(output_path)
print(f"✅ Excel文件已生成: {output_path}")
